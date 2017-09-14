#! /usr/env/bin python

"""
Load and clean up preliminary SnowEx'17 Pit sites (xlsx files)
Output csv of pit location, date, snow depth and bulk density 
"""
#First, fetch all xlsx files
#wget -r -A '*xlsx' ftp://ftp.nsidc.org/pub/projects/SnowEx/colorado_year1/Field_Books/

import sys
import os
import glob
from datetime import datetime
import re
import csv

import numpy as np
import openpyxl
from osgeo import osr

from pygeotools.lib import geolib

topdir = 'Pit_Output'
os.chdir(topdir)
#2017-02-06/PIT_L36/pit_20170206_L36.xlsx
xlsx_fn_list = glob.glob('*/*/*.xlsx')

#Define spatial reference for reprojection
utm13n_srs = osr.SpatialReference()
utm13n_srs.SetUTM(13, 1)
utm12n_srs = osr.SpatialReference()
utm12n_srs.SetUTM(12, 1)

print("Parsing %i files" % len(xlsx_fn_list))

outlist = []
for xlsx_fn in xlsx_fn_list:
    wb = openpyxl.load_workbook(xlsx_fn)
    ws = wb.worksheets[0]

    #Extract pit name from filename
    pitname = '_'.join(os.path.splitext(os.path.split(xlsx_fn)[-1])[0][12:].split('_')[1:])
    #pitname = ws['B6'].value

    #Extract systematic x and y coord
    #Initial x, y, zone 
    #print(ws['H4'].value, ws['H2'].value, ws['H6'].value)
    utmn = None
    if ws['H2'].value is not None:
        if '4826759' in str(ws['H2'].value):
            utmn = 4326759.0
        else:
            utmn = float(re.findall(r"[-+]?\d*\.\d+|\d+", str(ws['H2'].value).split('-')[-1])[0])
    utme = None
    if ws['H4'].value is not None:
        utme = float(re.findall(r"[-+]?\d*\.\d+|\d+", str(ws['H4'].value).split('-')[-1])[0])
    """
    utmzone = 13
    if ws['H6'].value is not None:
        if ws['H6'].value == 'RS':
            utmzone = 12
        elif ws['H6'].value == '*':
            utmzone = 13
        else:
            utmzone = int(ws['H6'].value.split()[0][0:2])
    """
    if utme is not None and utmn is not None:
        if utmn > 4400000:
            utmn /= 10.
        if utme > 4400000:
            utme /= 10.
        #These are sometimes swapped
        if utme > utmn:
            utmn, utme = utme, utmn

        #Determine UTM zone
        if int(str(utme)[0]) == 2:
            utmzone = 13
        elif int(str(utme)[0]) == 7:
            utmzone = 12
            #Transform coordinates to UTM 13N
            print("Transformng coordinates to UTM 13N")
            print(utme, utmn, 0, utm12n_srs, utm13n_srs)
            utme, utmn, dummy = geolib.cT_helper([utme,], [utmn,], [0,], utm12n_srs, utm13n_srs)
        #Some points are in lat/lon (39.01053, 108.187)
        elif int(str(utme)[0]) == 3: 
            print("Transformng coordinates to UTM 13N")
            utme, utmn, dummy = geolib.cT_helper(utmn, utme, 0, geolib.wgs_srs, utm13n_srs)
        else:
            continue

    #Extract depth
    depth = None
    if str(ws['F6'].value) is not None and str(ws['F6'].value)[0].isdigit():
        depth = float(str(ws['F6'].value).split('-')[-1].split()[0].split('cm')[0])
    elif str(ws['B10'].value) is not None and ws['B10'].value[0].isdigit():
        depth = float(ws['B10'].value)
    if depth is not None:
        depth /= 100.
    #There is one value that is 1100.0
    if depth > 1000:
        depth /= 10.
    #profile_depth = [float(x[0].value) for x in ws['B10:B33'] if x[0].value is not None and x[0].value[0].isdigit()]
    profile_depth = [float(x[0].value) for x in ws['B10:B33'] if x[0].value is not None and bool(re.match(r'^[0-9\.]*$', x[0].value))]
    profile_densityA = [float(x[0].value) for x in ws['E10:E33'] if x[0].value is not None and bool(re.match(r'^[0-9\.]*$', x[0].value))]
    profile_densityB = [float(x[0].value) for x in ws['F10:F33'] if x[0].value is not None and bool(re.match(r'^[0-9\.]*$', x[0].value))] 
    mean_density = np.hstack([profile_densityA, profile_densityB]).mean()

    #Extract datetime
    #pit_date = int(os.path.split(xlsx_fn)[-1].split('_')[1])
    #pit_time = map(int, ws['O6'].value.split(':'))
    #pit_date = os.path.split(xlsx_fn)[-1].split('_')[1]
    #pit_date = ws['M6'].value.split('/')
    pit_date = os.path.split(xlsx_fn)[-1][4:12]
    pit_time = None
    if ws['O6'].value is not None:
        pit_time = ws['O6'].value.split(':')
        pt = int(pit_time[0])
        #Attempt to deal with 12-hour vs 24-hour time
        if pt < 6:
            pt += 12
    if pit_time is not None:
        pit_dt_str = pit_date+'_'+pit_time[0]+pit_time[1]
        pit_dt = datetime.strptime(pit_dt_str, '%Y%m%d_%H%M') 
    else:
        pit_dt_str = pit_date
        pit_dt = datetime.strptime(pit_dt_str, '%Y%m%d') 

    out = [xlsx_fn, pitname, pit_dt, utme, utmn, depth, mean_density]
    out = tuple([np.nan if v is None else v for v in out])
    print(out)
    outlist.append(out)

csv_fn = 'snowex_pit_out.csv'
print("Writing out: %s" % csv_fn)

#This needs to be cleaned up 
hdr = 'file,pitname,datetime,x_utm13n,y_utm13n,depth_m,density_kgm3'.split(',')
fmt = '%s,%s,%s,%0.1f,%0.1f,%0.2f,%0.1f'

with open(csv_fn, "wb") as f:
    writer = csv.writer(f)
    writer.writerow(hdr)
    for row in outlist:
        writer.writerow((fmt % row).split(','))

#a = np.array(outlist)
#np.savetxt(csv_fn, a,fmt=fmt, delimiter=',', header=hdr) 
